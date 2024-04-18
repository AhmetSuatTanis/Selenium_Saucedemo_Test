from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
from constants.RegisterConstants import *
import json 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert


class Test_Register:
    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(Register_URL) #constants klasörü açarak içine değişkenler oluşturduk ordan çektik

    def teardown_method(self):
        self.driver.quit()

    
    def test_valid_register(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,userEmail)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        agreement1.click()
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        self.waitForElementAvailableForIFrame((By.XPATH,reCAPTHCHA_iframe_xpath))
        reCAPTHCHA=self.waitForElementClickable((By.CLASS_NAME, "recaptcha-checkbox-border"))
        reCAPTHCHA.click()
        #WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
        # WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
        self.driver.switch_to.default_content()
        sleep(15) #reCAPTHCHA için elle müdahele süresi
        continueButton=self.waitForElementVisible((By.CSS_SELECTOR,continueButton_CSS))
        continueButton.click()
        successRegisterMessage=self.waitForElementVisible((By.CSS_SELECTOR,successRegisterMessage_CSS))
        assert expedtedSuccessRegisterMessage in successRegisterMessage.text, f"'{expedtedSuccessRegisterMessage}' ifadesi mesaj içinde bulunamadı."

    
    
    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/bosBirakilanYerler.xlsx") #dosyanın nerde olduğunu gösterdik data klasöründe 
        sheet = excelFile["sheet1"] #sayfa değişkeni oluşturduk ve sayfayı söyledik
        rows = sheet.max_row #kaçıncı satıra kadar verim var onu söyledik
        data = []
        for i in range(2,rows): #parametreler 2.satırda olddan 2den başlattık, veri 4te bitiyor ama rows +1 yazıyoruz sondaki de dahil olsun diye
            username = sheet.cell(i,1).value #satırın 1.hücresi username'e gitsin. hücrenin içindeki değere ulaşmak için .value yazdık
            userLastname = sheet.cell(i,2).value
            password = sheet.cell(i,3).value #satırın 2.hücresi password'e gitsin
            passwordAgain = sheet.cell(i,4).value
            email = sheet.cell(i,5).value

            username = "" if username is None else username
            userLastname = "" if userLastname is None else userLastname
            password = "" if password is None else password
            passwordAgain = "" if passwordAgain is None else passwordAgain
            email = "" if email is None else email

            data.append((username,userLastname,password,passwordAgain,email)) 
        return data  #kullanılan noktaya bu datayı göndermek istediğimizi söylüyoruz
    
    @pytest.mark.parametrize("username,userLastname,password,passwordAgain,email",readInvalidDataFromExcel())
    def test_invalid_register(self,username,userLastname,password,passwordAgain,email):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,email)
        actions.perform()
        if email=="":  #email boş geçildiğinde alınan "doldurulması zorunlu alan" mesajı için if bloğu ile kontrol sağlandı.
            userEmailInput.send_keys("1") #direkt boş geçilince hata mesajı alınmadığı[Bug] için email içine karakter yazıp silindi.
            sleep(2)
            #üst kısımda yazılan Metini temizlemek için  
            userEmailInput.send_keys(Keys.CONTROL + "a")  # Tüm metni seç  #Keys metotlarını kullanmak için importunu ekledik.
            userEmailInput.send_keys(Keys.DELETE)         # Sil            #from selenium.webdriver.common.keys import Keys
            sleep(2)
            email_error_message_text=self.waitForElementVisible((By.XPATH,email_error_message_xpath))
            assert email_error_message==email_error_message_text.text, f"'{email_error_message}' mesajı görüntülenedi"     
        actions.send_keys_to_element(userPasswordInput,password)
        actions.send_keys_to_element(userPasswordAgainInput,passwordAgain)
        actions.perform()

        # Kayıt ol butonunun devre dışı olup olmadığını kontrol et
        #assert registerButton.get_attribute("disabled") == "true", "Kayıt ol butonu devre dışı değil."
        assert not registerButton.is_enabled(), "Kayıt ol butonu etkin durumda."

    
    def test_differentPasswords_errorMessage_register(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,userEmail)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword+"1")
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        agreement1.click()
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        self.waitForElementAvailableForIFrame((By.XPATH,reCAPTHCHA_iframe_xpath))
        reCAPTHCHA=self.waitForElementClickable((By.CLASS_NAME, "recaptcha-checkbox-border"))
        reCAPTHCHA.click()
        #WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
        # WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
        self.driver.switch_to.default_content()
        sleep(15) #reCAPTHCHA için elle müdahele süresi
        continueButton=self.waitForElementVisible((By.CSS_SELECTOR,continueButton_CSS))
        continueButton.click()
        sleep(1)
        popup_message = self.waitForElementVisible((By.XPATH,password_error_message_xpath))
        assert password_error_message in popup_message.text, f"'{password_error_message}' mesajı görüntülenemedi."

    def test_invalidEmail_errorMessage(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.send_keys_to_element(userEmailInput,invalidEmail)
        actions.perform()
        errorMessage=self.waitForElementVisible((By.XPATH,invalidEmail_errorMessage_xpath))
        assert invalidEmail_errorMessage ==errorMessage.text, f"'{invalidEmail_errorMessage}' ifadesi mesaj içinde bulunamadı."

    def test_registeredEmail_errorMessage(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,"ahmetsuattanis@gmail.com")
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        agreement1.click()
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        self.waitForElementAvailableForIFrame((By.XPATH,reCAPTHCHA_iframe_xpath))
        reCAPTHCHA=self.waitForElementClickable((By.CLASS_NAME, "recaptcha-checkbox-border"))
        reCAPTHCHA.click()
        self.driver.switch_to.default_content()
        sleep(15) #reCAPTHCHA için elle müdahele süresi bırakıldı
        continueButton=self.waitForElementVisible((By.CSS_SELECTOR,continueButton_CSS))
        continueButton.click()
        sleep(1)
        errorMessage=self.waitForElementVisible((By.XPATH,registeredEmail_errorMessage_xpath))
        assert registeredEmail_errorMessage in errorMessage.text , f"'{registeredEmail_errorMessage}' ifadesi mesaj içinde bulunamadı."

    def test_continueButton_passivityControl_withBlankAgreements(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,userEmail)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        #agreement1.click()  we did not click to this agreement for see continue button passive
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        self.waitForElementAvailableForIFrame((By.XPATH,reCAPTHCHA_iframe_xpath))
        reCAPTHCHA=self.waitForElementClickable((By.CLASS_NAME, "recaptcha-checkbox-border"))
        reCAPTHCHA.click()
        self.driver.switch_to.default_content()
        sleep(15) #reCAPTHCHA için elle müdahele süresi
        continueButton=self.waitForElementVisible((By.CSS_SELECTOR,continueButton_CSS))
        assert not continueButton.is_enabled(), "Devam Et butonu etkin durumda."

    def test_phoneNumberMaxMin_errorMessages(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,userEmail)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        agreement1.click()
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        sleep(2)
        phoneNumberInput.send_keys(Keys.BACK_SPACE)
        sleep(2)
        phoneNumberInput.send_keys(Keys.NUMPAD1)
        sleep(2)
        phoneNumberInput.send_keys(Keys.NUMPAD1)
        sleep(2)
        phoneNumberInput.send_keys(Keys.NUMPAD1)
        #EN AZ ve EN FAZLA KARAKTER HATASI  MESAJLARI ÇIKMIYOR , HOCAYA SOR bug mu diye

        maxErrorMessage=self.waitForElementVisible((By.XPATH,maxPhoneNumber_errorMessage_xpath))
        assert maxPhoneNumber_errorMessage==maxErrorMessage.text,f" '{maxErrorMessage}' ifadesi bulunamadı."
        phoneNumberInput.send_keys(Keys.BACK_SPACE)
        phoneNumberInput.send_keys(Keys.BACK_SPACE)
        phoneNumberInput.send_keys(Keys.BACK_SPACE)
        minErrorMessage=self.waitForElementVisible((By.XPATH,minPhoneNumber_errorMessage_xpath))
        assert minPhoneNumber_errorMessage==minErrorMessage.text,f" '{minErrorMessage}' ifadesi bulunamadı."

    #Bug #Bug #Bug önceden kullanılan numara ile tekrar üyelik açılabiliyor.
    def test_for_previouslyUsed_phoneNumber(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,userEmail)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        agreement1.click()
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        self.waitForElementAvailableForIFrame((By.XPATH,reCAPTHCHA_iframe_xpath))
        reCAPTHCHA=self.waitForElementClickable((By.CLASS_NAME, "recaptcha-checkbox-border"))
        reCAPTHCHA.click()
        self.driver.switch_to.default_content()
        sleep(15) #reCAPTHCHA için elle müdahele süresi
        continueButton=self.waitForElementVisible((By.CSS_SELECTOR,continueButton_CSS))
        continueButton.click()
        successRegisterMessage=self.waitForElementVisible((By.CSS_SELECTOR,successRegisterMessage_CSS))
        assert not expedtedSuccessRegisterMessage in successRegisterMessage.text, f"'{userPhoneNumber}' numarası ile daha önceden kayıt olunmasına rağmen tekrar kayıt olunabildi. Gereksinim dosyasına göre bu bir Bug'dır" 

    def test_reCAPTHCHA_errorMessage(self):
        userNameInput=self.waitForElementVisible((By.NAME,username_name))
        userLastNameInput=self.waitForElementVisible((By.NAME,userLastName_name))
        userEmailInput=self.waitForElementVisible((By.NAME,userEmail_name))
        userPasswordInput=self.waitForElementVisible((By.NAME,userPassword_name))
        userPasswordAgainInput=self.waitForElementVisible((By.NAME,userPasswordAgain_name))
        registerButton=self.waitForElementVisible((By.CSS_SELECTOR,registerButton_CSS))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userLastNameInput,userLastname)
        actions.send_keys_to_element(userEmailInput,userEmail)
        actions.send_keys_to_element(userPasswordInput,userPassword)
        actions.send_keys_to_element(userPasswordAgainInput,userPassword)
        actions.click(registerButton)
        actions.perform()
        agreement1=self.waitForElementVisible((By.NAME,agreement1_name))
        agreement1.click()
        agreement2=self.waitForElementVisible((By.NAME,agreement2_name))
        agreement2.click()
        agreement3=self.waitForElementVisible((By.NAME,agreement3_name))
        agreement3.click()
        agreement4=self.waitForElementVisible((By.NAME,agreement4_name))
        agreement4.click()
        phoneNumberInput=self.waitForElementVisible((By.ID,phoneNumberInput_id))
        phoneNumberInput.send_keys(userPhoneNumber)
        self.waitForElementAvailableForIFrame((By.XPATH,"//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]"))
        self.driver.find_element(By.CSS_SELECTOR, ".recaptcha-checkbox-border").click()
        sleep(58)
        assert self.waitForElementVisible(By.CSS_SELECTOR, reCAPTHCHA_errorMessage_CSS).text == "Verification expired. Check the checkbox again."
        # errorMessage=self.waitForElementPresence((By.XPATH,reCAPTHCHA_errorMessage_xpath))
        # assert reCAPTHCHA_errorMessage == errorMessage.text, f"'{reCAPTHCHA_errorMessage}' ifadesi mesaj içinde bulunamadı."
        #self.driver.switch_to.default_content()

        
        
        



    def waitForElementVisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
    
    def waitForElementAvailableForIFrame(self,locator,timeout=10):
        return WebDriverWait(self.driver, timeout).until(ec.frame_to_be_available_and_switch_to_it((locator)))
    
    def waitForElementClickable(self,locator,timeout=10):
        return WebDriverWait(self.driver, timeout).until(ec.element_to_be_clickable((locator)))
    
    
 