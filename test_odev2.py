from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
from constants.globalConstants import *


class Test_Odev:

    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL) #constants klasörü açarak içine değişkenler oluşturduk ordan çektik

    def teardown_method(self):
        self.driver.quit()
    
    
    def test_blank_login(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"")
        actions.send_keys_to_element(passwordInput,"")
        actions.click(loginButton)
        actions.perform() 
        #errorMessage1=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        errorMessage1=self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage1.text==errorMessage1_text
    
    def test_blank_password_login(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"Ahmet Suat Tanis")
        actions.send_keys_to_element(passwordInput,"")
        actions.click(loginButton)
        actions.perform() 
        #errorMessage2=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        errorMessage2=self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage2.text== errorMessage2_text
        

    def test_lockedUser_login(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"locked_out_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.click(loginButton)
        actions.perform() 
        #errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")))
        errorMessage3=self.waitForElementVisible((By.XPATH,errorMessage_xpath))
        assert errorMessage3.text==errorMessage3_text

    def test_valid_login(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        #baslik=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,baslik_xpath)))
        baslik=self.waitForElementVisible((By.XPATH,baslik_xpath))
        assert baslik.text== baslik_text
        listOfProducts=self.driver.find_elements(By.CSS_SELECTOR,listOfProducts_CSS)
        print(len(listOfProducts))
        assert len(listOfProducts)==listOfProducts_length


    def getData(): #bunu iptal ettik excelden çektik artık
        return [("ahmet","12345"),("naber","5341231"),("Ahmet Suat Tanis","secret_sauce")]

    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidCredentials.xlsx") #dosyanın nerde olduğunu gösterdik data klasöründe 
        sheet = excelFile["sheet1"] #sayfa değişkeni oluşturduk ve sayfayı söyledik
        rows = sheet.max_row #kaçıncı satıra kadar verim var onu söyledik
        data = []
        for i in range(2,rows+1): #parametreler 2.satırda olddan 2den başlattık, veri 4te bitiyor ama rows +1 yazıyoruz sondaki de dahil olsun diye
            username = sheet.cell(i,1).value #satırın 1.hücresi username'e gitsin. hücrenin içindeki değere ulaşmak için .value yazdık
            password = sheet.cell(i,2).value #satırın 2.hücresi password'e gitsin
            data.append((username,password)) 
        return data #kullanılan noktaya bu datayı göndermek istediğimizi söylüyoruz
    
    #artık pytest parametrize ile excel verilerimizi çağırabiliriz. get data örneğini siliyorum, parantez içine yeni defi yazıyorum


    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())  
    def test_invalid_login(self,username,password):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        errorMessage4=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        assert errorMessage4.text == errorMessage4_text

    
    def test_addToCartProduct(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        item1=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,item1_name)))
        item1.click()
        #cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,cartButton_CSS)))
        cartButton=self.waitForElementVisible((By.CSS_SELECTOR,cartButton_CSS))
        assert cartButton.text=="1"

    def test_removeProductFromCart(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        #item1=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,item1_name)))
        item2=self.waitForElementVisible((By.NAME,item2_name))
        item2.click()
        #item2=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,item2_CSS)))
        item3=self.waitForElementVisible((By.CSS_SELECTOR,item3_CSS))
        item3.click()
        #cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,cartButton_xpath)))
        cartButton=self.waitForElementVisible((By.XPATH,cartButton_xpath))
        cartButton.click()
        #removeButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"#remove-sauce-labs-bike-light")))
        removeButton=self.waitForElementVisible((By.CSS_SELECTOR,removeButton_CSS))
        removeButton.click()
        #cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,cartButton_xpath)))
        cartButton=self.waitForElementVisible((By.XPATH,cartButton_xpath))
        assert cartButton.text=="1"

    def test_checkoutProduct(self):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput=self.waitForElementVisible((By.ID,username_id))
        #passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput=self.waitForElementVisible((By.ID,password_id))
        #loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton=self.waitForElementVisible((By.ID,login_button_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        item4=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,item4_name)))
        item4.click()
        item5=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,item5_CSS)))
        item5.click()
        #cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,cartButton_xpath)))
        cartButton=self.waitForElementVisible((By.XPATH,cartButton_xpath))
        cartButton.click()
        #checkoutButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"checkout")))
        checkoutButton=self.waitForElementVisible((By.ID,checkoutButton_id))
        checkoutButton.click()
        #firstName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"first-name")))
        firstName=self.waitForElementVisible((By.ID, firstName_id))
        #lastName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"last-name")))
        lastName=self.waitForElementVisible((By.ID,lastName_id))
        #zipCode=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"postal-code")))
        zipCode=self.waitForElementVisible((By.ID,zipCode_id))
        #continueButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"continue")))
        continueButton=self.waitForElementVisible((By.ID,continueButton_id))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(firstName,"Yasemin")
        actions.send_keys_to_element(lastName,"Beyaz")
        actions.send_keys_to_element(zipCode,"34212")
        actions.perform()
        continueButton.click()
        #finishButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"finish")))
        finishButton=self.waitForElementVisible((By.ID,finishButton_id))
        actions.click(finishButton)
        actions.perform()
        #orderConfirmMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//h2[@class='complete-header']")))
        orderConfirmMessage=self.waitForElementVisible((By.XPATH,orderConfirmMessage_xpath))
        assert orderConfirmMessage.text== orderConfirmMessage_text

    def waitForElementVisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
    

    
    



        

    

        

    


    


    


    


 





    

        



